import tkinter as tk
from tkinter import ttk, messagebox
import requests
from bs4 import BeautifulSoup
import openpyxl
import os
import datetime

EXCEL_FILE = "mailing_list.xlsx"
LOG_FILE = "scraper.log"

def scrape_action(urls, username, password, status_box):
    added_count = 0
    skipped_count = 0

    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Email", "Phone", "Address", "Source URL"])
        wb.save(EXCEL_FILE)

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    emails = [row[1].value for row in ws.iter_rows(min_row=2)]

    for url in urls:
        try:
            response = requests.get(url.strip(), auth=(username, password))
            response.raise_for_status()
            soup = BeautifulSoup(response.text, "html.parser")

            name = soup.find("h1").get_text(strip=True) if soup.find("h1") else "N/A"
            email_tag = soup.find("a", href=lambda x: x and "mailto:" in x)
            email = email_tag.get_text(strip=True) if email_tag else "N/A"
            phone_tag = soup.find("span", class_="phone")
            phone = phone_tag.get_text(strip=True) if phone_tag else "N/A"
            address_tag = soup.find("div", class_="address")
            address = address_tag.get_text(strip=True) if address_tag else "N/A"

            record = [name, email, phone, address, url.strip()]

            if email not in emails:
                ws.append(record)
                emails.append(email)
                added_count += 1
                status_box.insert(tk.END, f"Added: {email} from {url}\n")
            else:
                skipped_count += 1
                status_box.insert(tk.END, f"Skipped duplicate: {email}\n")

            with open(LOG_FILE, "a") as log:
                log.write(f"{datetime.datetime.now()} | URL: {url} | Added: {email not in emails} | Email: {email}\n")

        except Exception as e:
            status_box.insert(tk.END, f"Error scraping {url}: {str(e)}\n")
            with open(LOG_FILE, "a") as log:
                log.write(f"{datetime.datetime.now()} | URL: {url} | ERROR: {str(e)}\n")

    wb.save(EXCEL_FILE)
    messagebox.showinfo("ScraperApp Pro", f"Scraping complete.\nAdded: {added_count}, Skipped: {skipped_count}")

def main():
    root = tk.Tk()
    root.title("ScraperApp Pro")
    root.geometry("600x500")

    style = ttk.Style()
    style.configure("TLabel", font=("Arial", 11))
    style.configure("TButton", font=("Arial", 11), padding=6)

    ttk.Label(root, text="Target URLs (one per line)").pack(pady=5)
    url_text = tk.Text(root, width=70, height=10)
    url_text.pack(pady=5)

    ttk.Label(root, text="Username").pack(pady=5)
    user_entry = ttk.Entry(root, width=50)
    user_entry.pack(pady=5)

    ttk.Label(root, text="Password").pack(pady=5)
    pass_entry = ttk.Entry(root, width=50, show="*")
    pass_entry.pack(pady=5)

    status_box = tk.Text(root, width=70, height=10, bg="#f0f0f0")
    status_box.pack(pady=10)

    def run_scrape():
        urls = url_text.get("1.0", tk.END).strip().splitlines()
        scrape_action(urls, user_entry.get(), pass_entry.get(), status_box)

    ttk.Button(root, text="Start Scraping", command=run_scrape).pack(pady=20)

    root.mainloop()

if __name__ == "__main__":
    main()
